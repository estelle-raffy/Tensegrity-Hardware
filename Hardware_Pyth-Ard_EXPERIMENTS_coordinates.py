## Builds on the Camera scripts 
## Sends error to Arduino ==> reads Serial monitor for all data ==> saves data in Excel
## Saves the coordinates (x, y) of the pink ball, with Timestamp, Elapsed time and Error distance...
# to facilitate comparison with first excel file 

from collections import deque
import numpy as np
import cv2
import imutils
import time
import serial
import os
import math
from datetime import datetime
from openpyxl import Workbook

# Define the lower and upper boundaries of the "pink" ball in HSL
pinkLower = (150, 100, 50)
pinkUpper = (180, 255, 200)

buffer_size = 32
pts = deque(maxlen=buffer_size)

# File setup
script_dir = os.path.dirname(os.path.abspath(__file__))
excel_file_path = os.path.join(script_dir, 'ExperimentsResults_Arduino-Python_tests.xlsx')
position_excel_path = os.path.join(script_dir, 'PositionTrackingLog.xlsx')

# Serial setup
try:
    ser = serial.Serial('COM3', 9600, timeout=1)
    time.sleep(2)
except serial.SerialException:
    print("[ERROR] Could not open serial port.")
    ser = None

# Video setup
print("[INFO] Starting video stream from external camera...")  # 0 is external, 1 is computer camera
vs = cv2.VideoCapture(1)

if not vs.isOpened():
    print("[ERROR] Could not open video stream.")
    exit()

time.sleep(2.0)

# Excel setup for Arduino data
wb = Workbook()
ws = wb.active
if ws.max_row == 1:
    ws.append([
        'Sample_Number', 'Motor', 'Local_Demand', 'Own_Voltage', 'Neighbour_Voltage',
        'Local_Error', 'Neighbour_Difference', 'Final_Actuation', 'New_Voltage',
        'New_Error', 'Local_Frustration', 'Weight', 'Error_python', 'Global_Frustration',
        'Global_Neighbour_Weight', 'Growth State'
    ])

# Excel setup for position tracking
pos_wb = Workbook()
pos_ws = pos_wb.active
if pos_ws.max_row == 1:
    pos_ws.append(['Timestamp', 'Elapsed_Time_s', 'X_Position', 'Y_Position', 'Error_Distance'])
start_time = time.time()

# Main loop
while True:

    # === 1️⃣ Receive all data from Arduino ===
    if ser.in_waiting > 0:
        print(" capture triggered")

        try:
            batch_data = []

            print(f"[DEBUG] Buffer content before reading: {ser.read(ser.in_waiting)}")
            time.sleep(0.1)

            while ser.in_waiting > 0:
                print("going to process ", ser.in_waiting)
                raw_line = ser.readline().decode('utf-8').strip()
                if raw_line:
                    batch_data.append(raw_line)
                else:
                    print("[WARN] Empty line received.")

                print(f"[DEBUG] Bytes waiting in buffer after read: {ser.in_waiting}")

            if batch_data:
                print(f"[DEBUG] Data received: {len(batch_data)} lines.")
                for line in batch_data:
                    print(f"[Sent to Excel] {line}:")
                    ws.append([line])  # Append each line to the Excel sheet
            else:
                print("[WARN] No data received.")

        except Exception as e:
            print(f"[ERROR] Failed to read data: {e}")
    else:
        print("[DEBUG] No data in buffer, waiting for new data...")
        time.sleep(0.1)

    # === 2️⃣ Camera tracking ===
    ret, frame = vs.read()
    if not ret:
        print("[ERROR] Failed to capture frame.")
        break

    frame = imutils.resize(frame, width=600)
    blurred = cv2.GaussianBlur(frame, (11, 11), 0)
    hsl = cv2.cvtColor(blurred, cv2.COLOR_BGR2HLS)
    mask = cv2.inRange(hsl, pinkLower, pinkUpper)
    mask = cv2.erode(mask, None, iterations=2)
    mask = cv2.dilate(mask, None, iterations=2)

    cnts = cv2.findContours(mask.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cnts = imutils.grab_contours(cnts)
    center = None

    frame_height, frame_width, _ = frame.shape
    target_x, target_y = frame_width // 2, frame_height // 2

    if len(cnts) > 0:
        c = max(cnts, key=cv2.contourArea)
        ((x, y), radius) = cv2.minEnclosingCircle(c)
        M = cv2.moments(c)
        center = (int(M["m10"] / M["m00"]), int(M["m01"] / M["m00"]))

        if radius > 10:
            cv2.circle(frame, (int(x), int(y)), int(radius), (0, 255, 255), 2)
            cv2.circle(frame, center, 5, (0, 0, 255), -1)
            pts.appendleft(center)

            error_distance = math.sqrt((target_x - center[0])**2 + (target_y - center[1])**2)
            print(f"[Computed Error] Distance: {error_distance:.2f} pixels")

            # === New logging of coordinates and time
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            elapsed_time = time.time() - start_time
            pos_ws.append([current_time, round(elapsed_time, 2), center[0], center[1], round(error_distance, 2)])

            # === Send error to Arduino
            if ser is not None:
                error_msg = f"{error_distance:.2f}\n"
                ser.write(error_msg.encode())
                ser.flush()
                print(f"[Sent to Arduino] {error_msg.strip()}")

    # === 3️⃣ Draw target cross ===
    cross_size = 20
    cross_color = (0, 255, 0)
    thickness = 2

    cv2.line(frame, (target_x - cross_size, target_y), (target_x + cross_size, target_y), cross_color, thickness)
    cv2.line(frame, (target_x, target_y - cross_size), (target_x, target_y + cross_size), cross_color, thickness)
    cv2.putText(frame, "Target", (target_x + 10, target_y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, cross_color, 2)

    cv2.imshow("Tracking", frame)

    if cv2.waitKey(1) & 0xFF == ord("q"):
        break

# === 4️⃣ Cleanup ===
vs.release()
cv2.destroyAllWindows()
if ser is not None:
    ser.close()
wb.save(excel_file_path)
pos_wb.save(position_excel_path)
print("[INFO] Experiment finished. Data saved to Excel.")
